#!/usr/bin/env python3 

# Excel Translator Program created by (reginapasela@gmail.com)
# This program use Deep-Translate as python module locally from: https://github.com/nidhaloff/deep-translator/

# Requirements:
# - deep_translate: 1.5.3
# - numpy: 1.20.2 
# - BeautifulSoup4: 4.9.3
# - requests: 2.25.1

# IMPORT LIBRARY AND MODULES
# @module numpy: for store data as array
#         openpyxl: to read/write excel
#         GoogleTranslator: for translate words
#         sys: to have commandline argument (sys.argv)
#         shutil: to have copy file command

import numpy as np
import openpyxl
from modules.deep_translator import GoogleTranslator
import sys
import re
from shutil import copyfile


# READ SOURCE EXCEL FILE LOCATION FROM USER VIA COMMANDLINE (sys.argv)
# @var src_loc: to store source excel file location
#      tgt_loc: to store output excel file location
# @type src_loc: str
#       tgt_loc: str

src_loc = sys.argv[1]
tgt_loc = sys.argv[2]


# COPY OUTPUT FILE FROM SOURCE FILE AND RENAME
# @var src_loc: source excel file location from user
#      tgt_loc: output excel file location from user
# @type src_loc: str
#       tgt_loc: str

copyfile(src_loc, tgt_loc)


# READ SOURCE LANGUAGE AND DESTINY LANGUAGE FROM USER
# @var src_lang: current excel language
#      tgt_lang: destionation excel language
# @type src_lang: str
#       tgt_lang: str

src_lang = str(input("Input your excel file source language = "))
tgt_lang = str(input("Input your excel file destination language = "))


# CREATING TRANSLATOR
# @var translator: to save GoogleTranslator language setting
# @param source: from user input in source variable
#        target: from user input in source variable

translator = GoogleTranslator(source = src_lang, target = tgt_lang)


# READ FROM SOURCE EXCEL FILE
# @var src_wb: to store source excel workbook file
#      tgt_wb: to store output excel workbook file

src_wb = openpyxl.load_workbook(src_loc)
tgt_wb = openpyxl.load_workbook(tgt_loc)


# PRINT ALL EXCEL SHEETS NAME IN src_wb

print(f"\nTotal sheets = {len(src_wb.sheetnames)}")
for i in range(len(src_wb.sheetnames)):
#   print format: "1. Sheet Name"     
    print(f"{i}.", src_wb.sheetnames[i])


# READ EXCEL SHEET INDEX NUMBER THAT WILL TRANSLATED FROM USER
# @var sheet_src: to store excel sheet index number from user
# @type sheet_src: str

try:
    sheet_idx = int(input("\nSelect sheet number you want to translate = "))
except:
    print("input number sheets only")


# READ WORK SHEET FOR SOURCE AND OUTPUT FILE
# @var src_sheet: variable to store source worksheet
#      tgt_sheet: variable to store output worksheet

src_sheet = src_wb[src_wb.sheetnames[sheet_idx]]
tgt_sheet = tgt_wb[tgt_wb.sheetnames[sheet_idx]]


# READ CELLS IN WORKSHEET AS NUMPY ARRAY
# @param dtype: to convert cells datatype to string

cells = np.array([[i.value for i in j] for j in src_sheet], dtype = str)


# DATA CLEANING
# 1. RESHAPE ARRAY TO 1 DIMENSION

cells = cells.reshape(1, cells.shape[0] * cells.shape[1])


# 2. FILTER None VALUE

cells = cells[cells != 'None']

# 3. FILTER FORMULA

# FILTER FORMULA
# This code will remove value in list if that list started with '=' 
r = re.compile(r'(?!=)')
cells = list(filter(r.match, cells))


# TRANSLATE ALL CELLS
# First we will save all translated word in a dictionary inside translations variable
# We are using .unique() to save time to translate duplicated words
# @var translations: variable to save all translated word
# @type translation: dictionary

translations = {}
unique_elements = np.unique([cells])
# create dictinary of translated words
for element in unique_elements:
#   first convert each element into string
    element = str(element)
#   check if element is digit then skip it, because we can't translate digit
    if element.isdigit():
        continue
#   add translated words into the dictionary
    translations[element] = translator.translate(element)


# REPLACE CELL VALUE WITH TRANSLATED WORDS IN translations dictionary
# We are using .value and .replace()

for r in range(1,tgt_sheet.max_row+1):
    for c in range(1,tgt_sheet.max_column+1):
        cell = str(tgt_sheet.cell(r,c).value)
        if cell in list(translations.keys()): 
            tgt_sheet.cell(r,c).value = cell.replace(cell, translations[cell])

# SAVE RESULT TO OUTPUT FILE
# We are using .save()

tgt_wb.save(tgt_loc)